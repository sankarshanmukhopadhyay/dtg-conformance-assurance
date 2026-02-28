#!/usr/bin/env python3
"""
Optional helper to export the source XLSX risk register to CSV.

CSV is the repo-first format. XLSX exists only as a convenience
for teams that maintain the risk register in spreadsheets.

Usage:
    python tools/export_xlsx_to_csv.py
"""
import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "risk" / "source" / "Risk Register.xlsx"
OUT = ROOT / "risk" / "exports" / "risk_assessment.csv"

CATEGORY_SHEETS_EXCLUDE = {"Explanation (using Gemini)", "Risk to Control Mapping"}

COLUMNS = [
    "Risk ID",
    "Category",
    "Risk Statement",
    "Cause",
    "Likelihood",
    "Impact",
    "Inherent Risk",
    "Controls",
    "Residual Risk",
]

def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source workbook: {SRC}")

    wb = load_workbook(SRC, data_only=True)
    rows_out = []
    for name in wb.sheetnames:
        if name in CATEGORY_SHEETS_EXCLUDE:
            continue
        ws = wb[name]
        # Expect header row at 1
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if not headers:
            continue
        idx = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            out = {col: "" for col in COLUMNS}
            out["Category"] = name
            for col in COLUMNS:
                if col in idx and idx[col] < len(row):
                    out[col] = row[idx[col]] if row[idx[col]] is not None else ""
            rows_out.append(out)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows_out)
    print(f"Wrote {len(rows_out)} rows to {OUT}")

if __name__ == "__main__":
    main()
